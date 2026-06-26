import base64
from pathlib import Path

from jinja2 import Environment, FileSystemLoader

_TEMPLATES_DIR = Path(__file__).parent.parent.parent / "templates"
_LOGO_PATH = Path(__file__).parent.parent.parent.parent / "frontend" / "public" / "udyogi-logo.png"


def _logo_b64() -> str:
    try:
        data = _LOGO_PATH.read_bytes()
        return "data:image/png;base64," + base64.b64encode(data).decode()
    except Exception:
        return ""

_ONES = [
    "", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
    "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen",
    "Seventeen", "Eighteen", "Nineteen",
]
_TENS = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]


def _words_below_hundred(n: int) -> str:
    if n < 20:
        return _ONES[n]
    tens = _TENS[n // 10]
    ones = _ONES[n % 10]
    return f"{tens} {ones}".strip() if ones else tens


def _words_below_thousand(n: int) -> str:
    if n < 100:
        return _words_below_hundred(n)
    h = _ONES[n // 100]
    rest = _words_below_hundred(n % 100)
    return f"{h} Hundred {rest}".strip() if rest else f"{h} Hundred"


def num_to_words(amount: int) -> str:
    """Convert an integer rupee amount to Indian-English words (e.g. 8493 → 'Eight Thousand Four Hundred Ninety Three Only')."""
    if amount == 0:
        return "Zero Only"
    parts = []
    crore = amount // 10_000_000
    if crore:
        parts.append(f"{_words_below_thousand(crore)} Crore")
        amount %= 10_000_000
    lakh = amount // 100_000
    if lakh:
        parts.append(f"{_words_below_thousand(lakh)} Lakh")
        amount %= 100_000
    thousand = amount // 1000
    if thousand:
        parts.append(f"{_words_below_thousand(thousand)} Thousand")
        amount %= 1000
    if amount:
        parts.append(_words_below_thousand(amount))
    return " ".join(parts) + " Only"


def _weasy_html():
    """Lazy WeasyPrint import with a clear error if system libs are missing."""
    try:
        from weasyprint import HTML as WeasyHTML  # needs pango/cairo at runtime
    except OSError as exc:
        raise RuntimeError(
            "WeasyPrint could not load system libraries (pango/cairo). "
            "Install them via brew (macOS) or apt (Debian) and retry. "
            f"Details: {exc}"
        ) from exc
    return WeasyHTML


def _render_html(template_name: str, context: dict) -> str:
    env = Environment(loader=FileSystemLoader(str(_TEMPLATES_DIR)))
    template = env.get_template(template_name)
    return template.render(**context, logo_b64=_logo_b64())


def generate_pdf(context: dict) -> bytes:
    """Render payslip_template.html with Jinja2 and convert to PDF bytes via WeasyPrint."""
    WeasyHTML = _weasy_html()
    html_string = _render_html("payslip_template.html", context)
    return WeasyHTML(string=html_string, base_url=str(_TEMPLATES_DIR)).write_pdf()


def generate_bulk_pdf(contexts: list[dict]) -> bytes:
    """Render one payslip per context and merge them into a single PDF (one per page)."""
    WeasyHTML = _weasy_html()
    docs = [
        WeasyHTML(string=_render_html("payslip_template.html", ctx),
                  base_url=str(_TEMPLATES_DIR)).render()
        for ctx in contexts
    ]
    pages = [page for doc in docs for page in doc.pages]
    return docs[0].copy(pages).write_pdf()


def generate_salary_sheet_xlsx(context: dict) -> bytes:
    """Build the salary-sheet payroll register as an .xlsx workbook (bytes).

    context: {entity_name, month_year, generated_on, rows:[...], totals:{...}}
    Each row dict carries the prorated, paid-amount figures from _build_response.
    """
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    headers = ["#", "Emp Code", "Name", "Designation", "Pay Days", "Basic", "HRA",
               "SPL", "CCA", "LTA", "Gross", "PF", "ESIC", "PT", "LD", "Loan",
               "Total Ded", "Net Pay", "PF (Empr)", "ESIC (Empr)"]
    # numeric data keys in column order, starting at the "Pay Days" column (E).
    num_keys = ["pay_days", "basic", "hra", "spl", "cca", "lta", "gross", "pf",
                "esic", "pt", "ld", "loan", "total_ded", "net", "pf_ern", "esic_ern"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Salary Sheet"

    ws.append([context["entity_name"]])
    ws.append([f"Salary Sheet — {context['month_year']}"])
    ws.append([f"Generated {context['generated_on']}"])
    ws.append([])
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"].font = Font(bold=True, size=11)
    ws["A3"].font = Font(italic=True, size=9, color="666666")

    header_row = ws.max_row + 1
    ws.append(headers)
    head_fill = PatternFill("solid", fgColor="E2E2E2")
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in ws[header_row]:
        c.font = Font(bold=True, size=9)
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for i, r in enumerate(context["rows"], start=1):
        ws.append([i, r["emp_code"], r["name"], r.get("designation") or ""]
                  + [r[k] for k in num_keys])

    t = context["totals"]
    # pay_days isn't summed (blank); the rest are column totals.
    total_row = ["", "", f"TOTAL — {len(context['rows'])} employees", "", ""]
    total_row += [t[k] for k in num_keys[1:]]
    ws.append(total_row)

    # number format + borders for the numeric block (cols E..T) across data + total
    first_data, last_row = header_row + 1, ws.max_row
    for row in ws.iter_rows(min_row=first_data, max_row=last_row, min_col=5, max_col=20):
        for c in row:
            if isinstance(c.value, (int, float)):
                c.number_format = "#,##0"
    for row in ws.iter_rows(min_row=header_row, max_row=last_row, min_col=1, max_col=20):
        for c in row:
            c.border = border
    for c in ws[last_row]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="EFEFEF")

    widths = [4, 12, 22, 18, 9] + [10] * 15
    from openpyxl.utils import get_column_letter
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Print setup: A3 landscape, fit to one page wide, repeat header row.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 8  # A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = f"{header_row}:{header_row}"

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
